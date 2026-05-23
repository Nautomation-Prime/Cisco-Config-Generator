from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook

from cisco_config_generator.orchestrator import Orchestrator


def _create_test_workbook(
    path: Path,
    *,
    model: str = "C9200-48P",
    uplink_module: str = "NM-4X",
    interface_rows: list[list[object]] | None = None,
) -> None:
    wb = Workbook()

    devices_ws = wb.active
    devices_ws.title = "Devices"
    devices_ws.append(
        [
            "Hostname",
            "Mgmt IP",
            "Mgmt VLAN",
            "Default Gateway",
            "Model",
            "Uplink Module",
            "Site",
            "Timezone",
            "Timezone Hours",
            "Timezone Minutes",
            "Mgmt Subnet",
        ]
    )
    devices_ws.append(
        [
            "SW-TEST-01",
            "10.0.10.2",
            10,
            "10.0.10.1",
            model,
            uplink_module,
            "HQ",
            "CET",
            1,
            0,
            "255.255.255.0",
        ]
    )

    global_ws = wb.create_sheet("Global Settings")
    global_ws.append(["Setting", "Value"])
    for row in [
        ("domain_name", "example.local"),
        ("snmp_auth_protocol", "SHA"),
        ("snmp_priv_protocol", "AES"),
        ("snmp_ro_group", "ROGROUP"),
        ("snmp_ro_user", "ro-user"),
        ("snmp_ro_auth_password", "authpass"),
        ("snmp_ro_priv_password", "privpass"),
        ("snmp_host", "10.0.1.20"),
        ("enable_secret", "TestSecret99!"),
        ("vty_acl", "ACL_VTY_ACCESS"),
        ("snmp_ro_acl", "ACL_SNMP_RO_ACCESS"),
        ("banner_motd", "TEST BANNER"),
    ]:
        global_ws.append(row)

    vlans_ws = wb.create_sheet("VLANs")
    vlans_ws.append(["VLAN ID", "VLAN Name", "Description"])
    vlans_ws.append([10, "MGMT", "Management VLAN"])
    vlans_ws.append([20, "DATA", "User Data VLAN"])

    interfaces_ws = wb.create_sheet("Interfaces")
    interfaces_ws.append(
        [
            "Device Name",
            "Interface Name",
            "Port Profile",
            "Description",
            "Access VLAN",
            "Voice VLAN",
            "Native VLAN",
            "Allowed VLANs",
            "Storm Control Broadcast",
            "Storm Control Multicast",
            "Port Channel No.",
        ]
    )
    rows = interface_rows or [
        [
            "SW-TEST-01",
            "GigabitEthernet1/0/1",
            "access-user",
            "Client Port",
            20,
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "SW-TEST-01",
            "TenGigabitEthernet1/1/1",
            "trunk-uplink-portchannel",
            "Uplink to Core",
            "",
            "",
            10,
            "10,20",
            "1.00 0.70",
            "1.00 0.70",
            1,
        ],
        [
            "SW-TEST-01",
            "TenGigabitEthernet1/1/2",
            "trunk-uplink-portchannel",
            "Uplink to Core",
            "",
            "",
            10,
            "10,20",
            "1.00 0.70",
            "1.00 0.70",
            1,
        ],
    ]
    for row in rows:
        interfaces_ws.append(row)

    acls_ws = wb.create_sheet("ACLs")
    acls_ws.append(["ACL Name", "Remark", "Action", "Network/Host", "Wildcard"])
    acls_ws.append(["ACL_VTY_ACCESS", "Management subnet", "permit", "10.0.0.0", "0.0.0.255"])
    acls_ws.append(["ACL_VTY_ACCESS", "Deny all others", "deny", "any", ""])
    acls_ws.append(["ACL_SNMP_RO_ACCESS", "NMS subnet", "permit", "10.0.1.0", "0.0.0.255"])
    acls_ws.append(["ACL_SNMP_RO_ACCESS", "Deny all others", "deny", "any", ""])

    features_ws = wb.create_sheet("Feature Selection")
    features_ws.append(["Feature", "Enabled", "Description"])
    features_ws.append(["base_config", "Yes", "Base switch config"])
    features_ws.append(["vlans", "Yes", "VLAN definitions"])
    features_ws.append(["interfaces", "Yes", "Interface config"])
    features_ws.append(["acls", "Yes", "ACL config"])

    wb.save(path)


class TestOrchestratorIntegration:
    def test_orchestrator_renders_full_config_from_workbook(self):
        repo_root = Path(__file__).resolve().parents[1]
        pack_path = repo_root / "packs" / "default"

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            workbook_path = tmp_path / "intent.xlsx"
            output_dir = tmp_path / "output"
            _create_test_workbook(workbook_path)

            orchestrator = Orchestrator(
                pack_path=pack_path,
                workbook_path=workbook_path,
                output_dir=output_dir,
            )

            written = orchestrator.run()

            assert len(written) == 1
            output_path = written[0]
            assert output_path.exists()
            assert output_path.name == "SW-TEST-01.cfg"

            config = output_path.read_text()

            assert "ip access-list standard ACL_VTY_ACCESS" in config
            assert "permit 10.0.0.0 0.0.0.255" in config
            assert "hostname SW-TEST-01" in config
            assert "clock timezone CET 1 0" in config
            assert "snmp-server group ROGROUP v3 priv read VIEWALL access ACL_SNMP_RO_ACCESS" in config
            assert "snmp-server host 10.0.1.20 version 3 priv ro-user" in config
            assert "access-class ACL_VTY_ACCESS in vrf-also" in config
            assert "vlan 20" in config
            assert "interface GigabitEthernet1/0/1" in config
            assert "switchport access vlan 20" in config
            assert "interface Port-channel1" in config
            assert config.count("interface Port-channel1") == 1
            assert "switchport trunk allowed vlan 10,20" in config
            assert config.count("storm-control broadcast level 1.00 0.70") == 4
            assert config.count("storm-control multicast level 1.00 0.70") == 4
            assert "storm-control broadcast level 0.10 0.07" not in config
            assert "storm-control multicast level 0.10 0.07" not in config
            assert "banner motd ^" in config
            assert config.index("hostname SW-TEST-01") < config.index("ip access-list standard ACL_VTY_ACCESS")

    def test_orchestrator_derives_missing_interfaces_from_selected_hardware(self):
        repo_root = Path(__file__).resolve().parents[1]
        pack_path = repo_root / "packs" / "default"

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            workbook_path = tmp_path / "intent.xlsx"
            output_dir = tmp_path / "output"
            _create_test_workbook(
                workbook_path,
                model="C9200-24P",
                uplink_module="NM-4X",
                interface_rows=[
                    [
                        "SW-TEST-01",
                        "GigabitEthernet1/0/1",
                        "access-user",
                        "Client Port",
                        20,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            )

            orchestrator = Orchestrator(
                pack_path=pack_path,
                workbook_path=workbook_path,
                output_dir=output_dir,
            )

            written = orchestrator.run()

            assert len(written) == 1
            config = written[0].read_text()

            assert "interface GigabitEthernet1/0/1" in config
            assert "switchport access vlan 20" in config
            assert config.count("default interface ") == 27
            assert "default interface GigabitEthernet1/0/2" in config
            assert "default interface GigabitEthernet1/0/24" in config
            assert "default interface TenGigabitEthernet1/1/1" in config
            assert "default interface TenGigabitEthernet1/1/4" in config
            assert "default interface GigabitEthernet1/0/25" not in config
