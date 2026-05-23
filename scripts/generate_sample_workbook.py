#!/usr/bin/env python3
"""
Generate assets/sample_intent.xlsx — a pre-filled example workbook that demonstrates
a realistic two-switch site deployment.

Run from the repo root:
    python scripts/generate_sample_workbook.py

The sample uses clearly fictional IP addresses and credentials so it can be committed
to source control without any risk of leaking real infrastructure details.
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl  # noqa: F401
except ImportError:
    print("openpyxl not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).parent.parent
TEMPLATE = REPO_ROOT / "assets" / "workbook_template.xlsx"
OUTPUT = REPO_ROOT / "assets" / "sample_intent.xlsx"

# ---------------------------------------------------------------------------
# Sample data
# ---------------------------------------------------------------------------

DEVICES = [
    # hostname, mgmt_ip, mgmt_vlan, gateway, model, uplink_module, site, tz, tz_h, tz_m, subnet
    ("SW-ACCESS-01", "10.10.10.11", 10, "10.10.10.1", "C9200-48P", "C9200-NM-4X", "Site-A", "GMT", 0, 0, "255.255.255.0"),
    ("SW-ACCESS-02", "10.10.10.12", 10, "10.10.10.1", "C9200-24P", "C9200-NM-4X", "Site-A", "GMT", 0, 0, "255.255.255.0"),
]

GLOBAL_SETTINGS = [
    ("domain_name",           "demo.nautomation.local"),
    ("ntp_servers",           "10.10.0.1, 10.10.0.2"),
    ("dns_servers",           "10.10.0.10, 10.10.0.11"),
    ("summer_time_config",    "BST recurring last Sun Mar 1:00 last Sun Oct 1:00"),
    ("aaa_group_name",        "TACACS_SERVERS"),
    ("tacacs_server_1",       "10.10.0.20"),
    ("tacacs_server_2",       "10.10.0.21"),
    ("tacacs_key",            "DemoTacacsKey99!"),
    ("aaa_fail_message",        "TACACS authentication failed - local account used."),
    ("snmp_auth_protocol",    "SHA"),
    ("snmp_priv_protocol",    "AES"),
    ("snmp_ro_group",         "SNMPv3_RO"),
    ("snmp_ro_user",          "snmp-ro-user"),
    ("snmp_ro_auth_password", "DemoRoAuth99!"),
    ("snmp_ro_priv_password", "DemoRoPriv99!"),
    ("snmp_rw_group",         "SNMPv3_RW"),
    ("snmp_rw_user",          "snmp-rw-user"),
    ("snmp_rw_auth_password", "DemoRwAuth99!"),
    ("snmp_rw_priv_password", "DemoRwPriv99!"),
    ("snmp_host",             "10.10.0.30"),
    ("snmp_location",         "Site-A — Server Room"),
    ("snmp_contact",          "netops@demo.nautomation.local"),
    ("vty_acl",               "ACL_VTY_ACCESS"),
    ("snmp_ro_acl",           "ACL_SNMP_RO"),
    ("snmp_rw_acl",           "ACL_SNMP_RW"),
    ("syslog_server",         "10.10.0.40"),
    ("banner_motd",           "AUTHORISED ACCESS ONLY — Nautomation Demo Site.\nDisconnect immediately if you are not authorised."),
    ("enable_secret",         "DemoEnableSecret99!"),
    ("local_username",        "netadmin"),
    ("local_password",        "DemoLocalPass99!"),
]

VLANS = [
    (10,  "MGMT",      "Management VLAN"),
    (20,  "DATA",      "User Data VLAN"),
    (30,  "VOICE",     "VoIP VLAN"),
    (40,  "WIRELESS",  "Wireless Client VLAN"),
    (50,  "PRINTERS",  "Printer VLAN"),
    (60,  "SERVERS",   "Server VLAN"),
    (70,  "CAMERAS",   "IP Camera VLAN"),
    (999, "UNUSED",    "Unused / Shutdown Ports"),
]

# Interfaces for SW-ACCESS-01 (48-port)
INTERFACES_SW01 = [
    # device, interface, profile, description, access_vlan, voice_vlan, native_vlan, allowed_vlans, sc_bc, sc_mc, pc_no
    ("SW-ACCESS-01", "GigabitEthernet1/0/1",      "access-voip",              "PC — Finance",        20, 30, "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/2",      "access-voip",              "PC — Finance",        20, 30, "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/3",      "access-voip",              "PC — HR",             20, 30, "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/4",      "access-voip",              "PC — HR",             20, 30, "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/5",      "access-voip",              "PC — IT Helpdesk",    20, 30, "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/6",      "access-voip",              "PC — IT Helpdesk",    20, 30, "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/7",      "access-printer",           "Printer — Floor 1",   50, "",  "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/8",      "access-printer",           "Printer — Floor 1",   50, "",  "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/9",      "access-ap-trunk",          "WAP — Ceiling MR46",  "",  "",  40, "20,30,40", "20.00 10.00", "5.00 2.00", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/10",     "access-ap-trunk",          "WAP — Ceiling MR46",  "",  "",  40, "20,30,40", "20.00 10.00", "5.00 2.00", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/11",     "access-video",             "IP Camera — Lobby",   70, "",  "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/12",     "access-video",             "IP Camera — Stairs",  70, "",  "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/13",     "access-control",           "Door Entry Reader",   20, "",  "", "", "", "", ""),
    ("SW-ACCESS-01", "GigabitEthernet1/0/48",     "access-server",            "OOB KVM",             60, "",  "", "", "", "", ""),
    ("SW-ACCESS-01", "TenGigabitEthernet1/1/1",   "trunk-uplink-portchannel", "Uplink to Core — Po1", "", "", 10, "10,20,30,40,50,60,70,999", "50.00 30.00", "20.00 10.00", 1),
    ("SW-ACCESS-01", "TenGigabitEthernet1/1/2",   "trunk-uplink-portchannel", "Uplink to Core — Po1", "", "", 10, "10,20,30,40,50,60,70,999", "50.00 30.00", "20.00 10.00", 1),
]

# Interfaces for SW-ACCESS-02 (24-port)
INTERFACES_SW02 = [
    ("SW-ACCESS-02", "GigabitEthernet1/0/1",      "access-voip",              "PC — Reception",      20, 30, "", "", "", "", ""),
    ("SW-ACCESS-02", "GigabitEthernet1/0/2",      "access-voip",              "PC — Meeting Rm A",   20, 30, "", "", "", "", ""),
    ("SW-ACCESS-02", "GigabitEthernet1/0/3",      "access-voip",              "PC — Meeting Rm B",   20, 30, "", "", "", "", ""),
    ("SW-ACCESS-02", "GigabitEthernet1/0/4",      "access-ap-trunk",          "WAP — Meeting Rm A",  "",  "",  40, "20,30,40", "20.00 10.00", "5.00 2.00", ""),
    ("SW-ACCESS-02", "GigabitEthernet1/0/5",      "access-printer",           "Printer — Floor 2",   50, "",  "", "", "", "", ""),
    ("SW-ACCESS-02", "GigabitEthernet1/0/6",      "access-video",             "IP Camera — Entrance",70, "",  "", "", "", "", ""),
    ("SW-ACCESS-02", "TenGigabitEthernet1/1/1",   "trunk-uplink-portchannel", "Uplink to Core — Po1", "", "", 10, "10,20,30,40,50,60,70,999", "50.00 30.00", "20.00 10.00", 1),
    ("SW-ACCESS-02", "TenGigabitEthernet1/1/2",   "trunk-uplink-portchannel", "Uplink to Core — Po1", "", "", 10, "10,20,30,40,50,60,70,999", "50.00 30.00", "20.00 10.00", 1),
]

ACLS = [
    # acl_name, remark, action, network, wildcard
    ("ACL_VTY_ACCESS", "Management hosts",           "permit", "10.10.0.0",  "0.0.0.255"),
    ("ACL_VTY_ACCESS", "IT admin workstations",     "permit", "10.10.1.0",  "0.0.0.255"),
    ("ACL_VTY_ACCESS", "Deny everything else",      "deny",   "any",        ""),
    ("ACL_SNMP_RO",    "NMS server",                "permit", "10.10.0.30", ""),
    ("ACL_SNMP_RO",    "Deny everything else",      "deny",   "any",        ""),
    ("ACL_SNMP_RW",    "NMS server",                "permit", "10.10.0.30", ""),
    ("ACL_SNMP_RW",    "Deny everything else",      "deny",   "any",        ""),
]

FEATURES = [
    ("base_config",  "Yes"),
    ("vlans",        "Yes"),
    ("interfaces",   "Yes"),
    ("acls",         "Yes"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_setting_row(ws, key: str) -> int | None:
    """Return the row number where column A matches key (case-insensitive)."""
    for row in ws.iter_rows(min_row=2):
        if row[0].value and str(row[0].value).strip().lower() == key.lower():
            return row[0].row
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def populate(wb) -> None:
    # ── Devices ────────────────────────────────────────────────────────────
    ws = wb["Devices"]
    # Clear example rows (rows 2 and 3 from template)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    for i, device in enumerate(DEVICES, start=2):
        for j, val in enumerate(device, start=1):
            ws.cell(row=i, column=j).value = val if val != "" else None

    # ── Global Settings ────────────────────────────────────────────────────
    ws = wb["Global Settings"]
    for key, value in GLOBAL_SETTINGS:
        r = _find_setting_row(ws, key)
        if r:
            ws.cell(row=r, column=2).value = value

    # ── VLANs ──────────────────────────────────────────────────────────────
    ws = wb["VLANs"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    for i, vlan in enumerate(VLANS, start=2):
        for j, val in enumerate(vlan, start=1):
            ws.cell(row=i, column=j).value = val

    # ── Interfaces ─────────────────────────────────────────────────────────
    ws = wb["Interfaces"]
    # Clear example rows (rows 3 onwards; row 2 is the instruction row)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    all_ifaces = INTERFACES_SW01 + INTERFACES_SW02
    for i, iface in enumerate(all_ifaces, start=3):
        for j, val in enumerate(iface, start=1):
            ws.cell(row=i, column=j).value = val if val != "" else None

    # ── ACLs ───────────────────────────────────────────────────────────────
    ws = wb["ACLs"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    for i, acl in enumerate(ACLS, start=2):
        for j, val in enumerate(acl, start=1):
            ws.cell(row=i, column=j).value = val if val != "" else None

    # ── Feature Selection ──────────────────────────────────────────────────
    ws = wb["Feature Selection"]
    for key, value in FEATURES:
        r = _find_setting_row(ws, key)
        if r:
            ws.cell(row=r, column=2).value = value


def main() -> None:
    if not TEMPLATE.exists():
        print(f"Template not found: {TEMPLATE}")
        print("Run:  python scripts/generate_workbook.py  first.")
        sys.exit(1)

    wb = load_workbook(TEMPLATE)
    populate(wb)
    wb.save(OUTPUT)
    print(f"Sample workbook generated: {OUTPUT}")


if __name__ == "__main__":
    main()
