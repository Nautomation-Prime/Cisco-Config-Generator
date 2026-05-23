#!/usr/bin/env python3
"""
Generate the intent workbook template: assets/workbook_template.xlsx

Run from the repo root:
    python scripts/generate_workbook.py
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install -r requirements.txt")
    sys.exit(1)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

REPO_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(REPO_ROOT))

import yaml

HARDWARE_CATALOG = yaml.safe_load((REPO_ROOT / "packs/default/hardware_catalog.yaml").read_text())
PORT_PROFILES = yaml.safe_load((REPO_ROOT / "packs/default/port_profiles.yaml").read_text())

MODELS = list(HARDWARE_CATALOG["switch_models"].keys())
UPLINK_MODULES = list(HARDWARE_CATALOG["uplink_modules"].keys())
PROFILES = list(PORT_PROFILES["profiles"].keys())

# --- Styles ---
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
ROW_FILL_ALT = PatternFill("solid", fgColor="DCE6F1")
ROW_FILL_NORM = PatternFill("solid", fgColor="FFFFFF")
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
_BORDER_SIDE = Side(style="thin", color="B8CCE4")

# --- Tab colours ---
TAB_GREY = "595959"
TAB_BLUE = "1F3864"
TAB_GREEN = "375623"
TAB_ORANGE = "C55A11"
TAB_PURPLE = "7030A0"
TAB_TEAL = "0070C0"
TAB_RED = "C00000"


def _border() -> Border:
    return Border(left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE)


def style_header(ws, headers: list[str], widths: list[int] | None = None, row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = (widths[col_idx - 1] if widths else 22)
    ws.row_dimensions[row].height = 25


def apply_row_style(ws, row: int, num_cols: int, alt: bool = False) -> None:
    fill = ROW_FILL_ALT if alt else ROW_FILL_NORM
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = NORMAL_FONT
        cell.border = _border()
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 18


def add_list_validation(ws, col: int, formula: str, start_row: int = 2, end_row: int = 300) -> None:
    col_letter = get_column_letter(col)
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        error="Please select a valid value from the dropdown.",
        errorTitle="Invalid Value",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"


def add_named_ranges(wb: Workbook) -> None:
    """Named ranges enable cross-sheet dropdown references in Excel."""
    wb.defined_names["VLANIDs"] = DefinedName("VLANIDs", attr_text="VLANs!$A$2:$A$300")
    wb.defined_names["DeviceNames"] = DefinedName("DeviceNames", attr_text="Devices!$A$2:$A$300")


# -----------------------------------------------------------------------
# Instructions sheet (inserted at position 0 — first tab)
# -----------------------------------------------------------------------
def create_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = TAB_GREY
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 65

    # Title
    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1)
    title.value = "Cisco Config Generator — Intent Workbook"
    title.font = Font(name="Calibri", size=16, bold=True, color="1F3864")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:B2")
    sub = ws.cell(row=2, column=1)
    sub.value = "Nautomation Prime  |  Fill in all sheets then run the tool to generate configs."
    sub.font = Font(name="Calibri", size=11, italic=True, color="595959")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8  # spacer

    # Sheet guide table
    guide = [
        ("Sheet", "What to fill in"),
        ("Devices", "One row per switch. Select Model and Uplink Module from the dropdowns. Those hardware choices drive the full physical interface inventory used by the generator. Each device needs a unique Hostname, Management IP, VLAN, Default Gateway, timezone label, and timezone hour/minute offsets."),
        ("Global Settings", "Settings shared across all devices — NTP/DNS servers, TACACS+ servers and key, SNMPv3 RO/RW credentials, SNMP host/location/contact, ACL names for VTY and SNMP access (bodies defined in the ACLs sheet), syslog server, banner, enable secret, and local admin credentials."),
        ("VLANs", "Define all VLANs for the site. VLAN IDs entered here automatically appear in the Access VLAN and Voice VLAN dropdowns on the Interfaces sheet."),
        ("Interfaces", "Add rows only for ports that need explicit config. The generator derives the full interface list from each device's selected Model and Uplink Module; any omitted ports default to the unused profile and render as shutdown ports. Access/Voice/Native VLAN dropdowns are linked to the VLANs sheet. Trunk and AP-trunk ports use Native VLAN, optional Allowed VLANs, optional Storm Control Broadcast/Multicast overrides, and Port Channel No. where required; access ports use Access/Voice VLAN."),
        ("ACLs", "Define IP access control lists used for VTY and SNMP access restriction. One row per ACL entry: ACL Name, optional Remark, Action (permit/deny), Network/Host, and optional Wildcard mask. ACL names must match those set in Global Settings (vty_acl, snmp_ro_acl, snmp_rw_acl)."),
        ("Feature Selection", "Toggle which config sections to generate (Yes/No). Useful if you only need to regenerate interface configs, for example."),
    ]
    for i, (sheet, desc) in enumerate(guide):
        row = i + 4
        a = ws.cell(row=row, column=1, value=sheet)
        b = ws.cell(row=row, column=2, value=desc)
        if i == 0:
            a.font = HEADER_FONT
            b.font = HEADER_FONT
            a.fill = HEADER_FILL
            b.fill = HEADER_FILL
            a.alignment = Alignment(horizontal="center", vertical="center")
            b.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 22
        else:
            fill = ROW_FILL_ALT if i % 2 == 0 else ROW_FILL_NORM
            a.font = BOLD_FONT
            b.font = NORMAL_FONT
            a.fill = fill
            b.fill = fill
            b.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 38

    ws.row_dimensions[len(guide) + 5].height = 10  # spacer

    # Tips section
    tips_row = len(guide) + 6
    ws.merge_cells(f"A{tips_row}:B{tips_row}")
    tip_hdr = ws.cell(row=tips_row, column=1, value="Tips")
    tip_hdr.font = Font(name="Calibri", size=12, bold=True, color="1F3864")
    ws.row_dimensions[tips_row].height = 22

    tips = [
        "• Fill in the VLANs sheet before the Interfaces sheet so VLAN dropdowns are populated.",
        "• Fill in the Devices sheet before the Interfaces sheet so the Device Name dropdown works and the generator can derive the full port inventory.",
        "• You only need to add rows on the Interfaces sheet for ports that are in use or need custom settings. Omitted ports are auto-generated as unused from the selected model and uplink module.",
        "• To add a new switch model or port profile, edit the YAML files in packs/default/ and re-run scripts/generate_workbook.py.",
        "• Save your completed workbook with a descriptive name (e.g. site-london-2024.xlsx).",
        "• Generated configs are written to the output\\ folder — one .cfg file per device.",
    ]
    for j, tip in enumerate(tips):
        r = tips_row + 1 + j
        ws.merge_cells(f"A{r}:B{r}")
        cell = ws.cell(row=r, column=1, value=tip)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 18


# -----------------------------------------------------------------------
# Sheet 1 — Devices
# -----------------------------------------------------------------------
def create_devices_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Devices"
    ws.sheet_properties.tabColor = TAB_BLUE
    ws.freeze_panes = "A2"

    headers = ["Hostname", "Mgmt IP", "Mgmt VLAN", "Default Gateway",
               "Model", "Uplink Module", "Site", "Timezone", "Timezone Hours", "Timezone Minutes", "Mgmt Subnet"]
    widths = [22, 18, 14, 20, 22, 20, 20, 14, 16, 18, 18]
    style_header(ws, headers, widths=widths)

    add_list_validation(ws, col=5, formula='"' + ",".join(MODELS) + '"')
    add_list_validation(ws, col=6, formula='"' + ",".join(UPLINK_MODULES) + '"')

    example_rows = [
        ("SW-OFFICE-01", "10.0.10.2", 10, "10.0.10.1", MODELS[0], UPLINK_MODULES[0], "Head Office", "GMT", 0, 0, "255.255.255.0"),
        ("SW-OFFICE-02", "10.0.10.3", 10, "10.0.10.1", MODELS[1], UPLINK_MODULES[0], "Head Office", "CET", 1, 0, "255.255.255.0"),
    ]
    for i, row_data in enumerate(example_rows):
        ws.append(row_data)
        apply_row_style(ws, ws.max_row, len(headers), alt=(i % 2 == 1))


# -----------------------------------------------------------------------
# Sheet 2 — Global Settings
# -----------------------------------------------------------------------
def create_global_settings_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Global Settings")
    ws.sheet_properties.tabColor = TAB_GREEN
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    for col, val in enumerate(["Setting", "Value"], start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = val
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[1].height = 25

    rows = [
        ("domain_name",             "example.local"),
        ("ntp_servers",             "10.0.0.1, 10.0.0.2"),
        ("dns_servers",             "8.8.8.8, 8.8.4.4"),
        ("summer_time_config",      "BST recurring last Sun Mar 1:00 last Sun Oct 1:00"),
        ("aaa_group_name",          "AAA_NOC"),
        ("tacacs_server_1",         ""),
        ("tacacs_server_2",         ""),
        ("tacacs_key",              ""),
        ("aaa_fail_message",        ""),
        ("snmp_auth_protocol",      "SHA"),
        ("snmp_priv_protocol",      "AES"),
        ("snmp_ro_group",           "ROGROUP"),
        ("snmp_ro_user",            "ROBOOK"),
        ("snmp_ro_auth_password",   "changeme_auth"),
        ("snmp_ro_priv_password",   "changeme_priv"),
        ("snmp_rw_group",           "RWGROUP"),
        ("snmp_rw_user",            "RWBOOK"),
        ("snmp_rw_auth_password",   "changeme_auth"),
        ("snmp_rw_priv_password",   "changeme_priv"),
        ("snmp_host",               ""),
        ("snmp_location",           ""),
        ("snmp_contact",            ""),
        ("vty_acl",                 "ACL_VTY_ACCESS"),
        ("snmp_ro_acl",             "ACL_SNMP_RO_ACCESS"),
        ("snmp_rw_acl",             "ACL_SNMP_RW_ACCESS"),
        ("syslog_server",           ""),
        ("banner_motd",             "AUTHORISED ACCESS ONLY. Disconnect immediately if not authorised."),
        ("enable_secret",           "changeme"),
        ("local_username",          "admin"),
        ("local_password",          "changeme"),
    ]
    AUTH_PROTOCOL_KEYS = {"snmp_auth_protocol"}
    PRIV_PROTOCOL_KEYS = {"snmp_priv_protocol"}

    for i, row_data in enumerate(rows):
        ws.append(row_data)
        r = ws.max_row
        fill = ROW_FILL_ALT if i % 2 == 0 else ROW_FILL_NORM
        for col in range(1, 3):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.font = NORMAL_FONT
            cell.border = _border()
            cell.alignment = Alignment(vertical="center")
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.row_dimensions[r].height = 18
        key = row_data[0]
        if key in AUTH_PROTOCOL_KEYS:
            dv = DataValidation(type="list", formula1='"SHA,MD5"', allow_blank=False, showDropDown=False)
            ws.add_data_validation(dv)
            dv.sqref = f"B{r}"
        elif key in PRIV_PROTOCOL_KEYS:
            dv = DataValidation(type="list", formula1='"AES,DES"', allow_blank=False, showDropDown=False)
            ws.add_data_validation(dv)
            dv.sqref = f"B{r}"


# -----------------------------------------------------------------------
# Sheet 3 — VLANs
# -----------------------------------------------------------------------
def create_vlans_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("VLANs")
    ws.sheet_properties.tabColor = TAB_ORANGE
    ws.freeze_panes = "A2"

    style_header(ws, ["VLAN ID", "VLAN Name", "Description"], widths=[12, 22, 45])

    example_vlans = [
        (10, "MGMT", "Management VLAN"),
        (20, "DATA", "User Data VLAN"),
        (30, "VOICE", "VoIP VLAN"),
        (40, "WIRELESS", "Wireless VLAN"),
        (999, "UNUSED", "Unused ports VLAN"),
    ]
    for i, vlan in enumerate(example_vlans):
        ws.append(vlan)
        apply_row_style(ws, ws.max_row, 3, alt=(i % 2 == 1))


# -----------------------------------------------------------------------
# Sheet 4 — Interfaces
# -----------------------------------------------------------------------
def create_interfaces_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Interfaces")
    ws.sheet_properties.tabColor = TAB_PURPLE
    ws.freeze_panes = "A2"

    headers = ["Device Name", "Interface Name", "Port Profile",
               "Description", "Access VLAN", "Voice VLAN", "Native VLAN", "Allowed VLANs", "Storm Control Broadcast", "Storm Control Multicast", "Port Channel No."]
    widths = [22, 32, 24, 40, 14, 14, 14, 22, 24, 24, 16]
    style_header(ws, headers, widths=widths)

    # Device Name — cross-sheet dropdown from Devices!$A (named range)
    add_list_validation(ws, col=1, formula="DeviceNames")
    # Port Profile — inline list from pack YAML
    add_list_validation(ws, col=3, formula='"' + ",".join(PROFILES) + '"')
    # Access VLAN / Voice VLAN / Native VLAN — cross-sheet dropdown from VLANs!$A (named range)
    add_list_validation(ws, col=5, formula="VLANIDs")
    add_list_validation(ws, col=6, formula="VLANIDs")
    add_list_validation(ws, col=7, formula="VLANIDs")

    example_rows = [
        ("SW-OFFICE-01", "GigabitEthernet1/0/1",    "access-voip",              "CLIENT: PC/Phone - Desk 1",   20, 30,  "",  "",            "",          "",          ""),
        ("SW-OFFICE-01", "GigabitEthernet1/0/2",    "access-printer",           "PRINTER: Floor 1",            20, "",  "",  "",            "",          "",          ""),
        ("SW-OFFICE-01", "GigabitEthernet1/0/3",    "access-ap-trunk",          "WAP - Lobby",                 "", "",  40,  "10,20,40",    "1.00 0.70", "1.00 0.70", ""),
        ("SW-OFFICE-01", "TenGigabitEthernet1/1/1", "trunk-uplink-portchannel", "L:Po1  R:Po1 Uplink to Core", "", "",  99,  "10,20,40,99", "1.00 0.70", "1.00 0.70", 1),
        ("SW-OFFICE-01", "TenGigabitEthernet1/1/2", "trunk-uplink-portchannel", "L:Po1  R:Po1 Uplink to Core", "", "",  99,  "10,20,40,99", "1.00 0.70", "1.00 0.70", 1),
    ]
    for i, row_data in enumerate(example_rows):
        ws.append(row_data)
        apply_row_style(ws, ws.max_row, len(headers), alt=(i % 2 == 1))


# -----------------------------------------------------------------------
# Sheet 5 — ACLs
# -----------------------------------------------------------------------
def create_acls_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("ACLs")
    ws.sheet_properties.tabColor = TAB_RED
    ws.freeze_panes = "A2"

    headers = ["ACL Name", "Remark", "Action", "Network/Host", "Wildcard"]
    widths = [28, 40, 12, 22, 18]
    style_header(ws, headers, widths=widths)

    add_list_validation(ws, col=3, formula='"permit,deny"', start_row=2, end_row=300)

    example_rows = [
        ("ACL_VTY_ACCESS",      "Management network",   "permit", "10.0.0.0",  "0.0.0.255"),
        ("ACL_VTY_ACCESS",      "Jump host",            "permit", "10.0.1.10", ""),
        ("ACL_VTY_ACCESS",      "Deny all others",      "deny",   "any",       ""),
        ("ACL_SNMP_RO_ACCESS",  "NMS server",           "permit", "10.0.1.20", ""),
        ("ACL_SNMP_RO_ACCESS",  "Deny all others",      "deny",   "any",       ""),
        ("ACL_SNMP_RW_ACCESS",  "NMS server",           "permit", "10.0.1.20", ""),
        ("ACL_SNMP_RW_ACCESS",  "Deny all others",      "deny",   "any",       ""),
    ]
    for i, row_data in enumerate(example_rows):
        ws.append(row_data)
        apply_row_style(ws, ws.max_row, len(headers), alt=(i % 2 == 1))


# -----------------------------------------------------------------------
# Sheet 6 — Feature Selection
# -----------------------------------------------------------------------
def create_feature_selection_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Feature Selection")
    ws.sheet_properties.tabColor = TAB_TEAL
    ws.freeze_panes = "A2"

    style_header(ws, ["Feature", "Enabled", "Description"], widths=[22, 12, 55])
    add_list_validation(ws, col=2, formula='"Yes,No"', start_row=2, end_row=50)

    features = [
        ("base_config", "Yes", "Base switch config — hostname, AAA, NTP, SNMP, STP, SSH hardening"),
        ("vlans",       "Yes", "VLAN definitions"),
        ("interfaces",  "Yes", "Interface config — access, trunk, and unused ports"),
        ("acls",        "Yes", "IP access control lists — VTY and SNMP access restriction"),
    ]
    for i, row_data in enumerate(features):
        ws.append(row_data)
        apply_row_style(ws, ws.max_row, 3, alt=(i % 2 == 1))


def main() -> None:
    output_path = REPO_ROOT / "assets" / "workbook_template.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Create sheets (Devices uses wb.active — the default blank sheet)
    create_devices_sheet(wb)
    create_global_settings_sheet(wb)
    create_vlans_sheet(wb)
    create_interfaces_sheet(wb)
    create_acls_sheet(wb)
    create_feature_selection_sheet(wb)

    # Named ranges must exist before cross-sheet dropdowns are usable
    add_named_ranges(wb)

    # Insert Instructions as the first tab (shifts all others right)
    create_instructions_sheet(wb)

    try:
        wb.save(str(output_path))
    except PermissionError as exc:
        print(
            f"Could not write workbook template to {output_path} because the file is open. "
            "Close the workbook in Excel and run the generator again."
        )
        raise SystemExit(1) from exc

    print(f"Workbook template generated: {output_path}")


if __name__ == "__main__":
    main()
