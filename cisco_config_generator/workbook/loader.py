from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cisco_config_generator.workbook.models import (
    Device, VLAN, Interface, GlobalSettings, FeatureSelection, Intent, ACLEntry
)


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    val = ws.cell(row=row, column=col).value
    return val if val is not None else ""


def _list_field(raw: Any) -> list[str]:
    if not raw:
        return []
    return [s.strip() for s in str(raw).split(",") if s.strip()]


def _header_map(ws: Worksheet) -> dict[str, int]:
    """Return {header_lower: col_index} from row 1 of a worksheet."""
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip().lower().replace(" ", "_")] = col
    return headers


def _validate_required_headers(headers: dict[str, int], required: list[str], sheet_name: str) -> None:
    """Raise ValueError listing any required column headers absent from the sheet."""
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing required column(s): "
            + ", ".join(f"'{h}'" for h in missing)
            + f". Found: {list(headers.keys())}"
        )


def _parse_int_cell(raw: Any, field: str, row: int, sheet: str, default: int | None = None) -> int | None:
    """Parse a cell value as int, raising ValueError with row/field context if it fails."""
    if not str(raw).strip():
        return default
    try:
        return int(raw)
    except (ValueError, TypeError):
        raise ValueError(
            f"Sheet '{sheet}', row {row}, column '{field}': expected integer, got '{raw}'."
        )


def _load_devices(ws: Worksheet) -> list[Device]:
    headers = _header_map(ws)
    _validate_required_headers(
        headers,
        ["hostname", "mgmt_ip", "mgmt_vlan", "default_gateway", "model", "uplink_module"],
        ws.title,
    )
    sheet = ws.title
    devices: list[Device] = []
    for row in range(2, ws.max_row + 1):
        hostname = _cell_value(ws, row, headers.get("hostname", 1))
        if not hostname:
            continue
        timezone_hours_col = headers.get("timezone_hours")
        timezone_minutes_col = headers.get("timezone_minutes")
        timezone_hours_raw = _cell_value(ws, row, timezone_hours_col) if timezone_hours_col else ""
        timezone_minutes_raw = _cell_value(ws, row, timezone_minutes_col) if timezone_minutes_col else ""
        devices.append(Device(
            hostname=str(hostname).strip(),
            mgmt_ip=str(_cell_value(ws, row, headers.get("mgmt_ip", 2))).strip(),
            mgmt_vlan=_parse_int_cell(_cell_value(ws, row, headers.get("mgmt_vlan", 3)), "Mgmt VLAN", row, sheet, default=1),
            default_gateway=str(_cell_value(ws, row, headers.get("default_gateway", 4))).strip(),
            model=str(_cell_value(ws, row, headers.get("model", 5))).strip(),
            uplink_module=str(_cell_value(ws, row, headers.get("uplink_module", 6))).strip(),
            site=str(_cell_value(ws, row, headers.get("site", 7))).strip(),
            timezone=str(_cell_value(ws, row, headers.get("timezone", 8)) or "GMT").strip(),
            timezone_hours_offset=_parse_int_cell(timezone_hours_raw, "Timezone Hours", row, sheet, default=0),
            timezone_minutes_offset=_parse_int_cell(timezone_minutes_raw, "Timezone Minutes", row, sheet, default=0),
            mgmt_subnet=str(_cell_value(ws, row, headers.get("mgmt_subnet", 9)) or "255.255.255.0").strip(),
        ))
    return devices


def _load_vlans(ws: Worksheet) -> list[VLAN]:
    headers = _header_map(ws)
    _validate_required_headers(headers, ["vlan_id", "vlan_name"], ws.title)
    sheet = ws.title
    vlans: list[VLAN] = []
    for row in range(2, ws.max_row + 1):
        vlan_id = _cell_value(ws, row, headers.get("vlan_id", 1))
        if not vlan_id:
            continue
        vlans.append(VLAN(
            vlan_id=_parse_int_cell(vlan_id, "VLAN ID", row, sheet) or 0,
            vlan_name=str(_cell_value(ws, row, headers.get("vlan_name", 2))).strip(),
            description=str(_cell_value(ws, row, headers.get("description", 3))).strip(),
        ))
    return vlans


def _load_interfaces(ws: Worksheet, port_profiles: dict[str, Any]) -> list[Interface]:
    headers = _header_map(ws)
    _validate_required_headers(headers, ["device_name", "interface_name", "port_profile"], ws.title)
    sheet = ws.title
    interfaces: list[Interface] = []
    for row in range(2, ws.max_row + 1):
        device_name = _cell_value(ws, row, headers.get("device_name", 1))
        if not device_name:
            continue
        port_profile = str(_cell_value(ws, row, headers.get("port_profile", 3))).strip()
        template_hint = ""
        qos_trust_dscp = False
        if port_profile and port_profiles:
            profile_data = port_profiles.get(port_profile, {})
            template_hint = profile_data.get("template_hint", "")
            qos_trust_dscp = bool(profile_data.get("qos_trust_dscp", False))
        access_vlan_raw = _cell_value(ws, row, headers.get("access_vlan", 5))
        voice_vlan_raw = _cell_value(ws, row, headers.get("voice_vlan", 6))
        native_vlan_raw = _cell_value(ws, row, headers.get("native_vlan", 7))
        allowed_vlans_col = headers.get("allowed_vlans")
        storm_control_broadcast_col = headers.get("storm_control_broadcast")
        storm_control_multicast_col = headers.get("storm_control_multicast")
        port_channel_col = headers.get("port_channel_no.")
        allowed_vlans_raw = _cell_value(ws, row, allowed_vlans_col) if allowed_vlans_col else ""
        storm_control_broadcast_raw = _cell_value(ws, row, storm_control_broadcast_col) if storm_control_broadcast_col else ""
        storm_control_multicast_raw = _cell_value(ws, row, storm_control_multicast_col) if storm_control_multicast_col else ""
        port_channel_raw = _cell_value(ws, row, port_channel_col) if port_channel_col else ""
        interfaces.append(Interface(
            device_name=str(device_name).strip(),
            interface_name=str(_cell_value(ws, row, headers.get("interface_name", 2))).strip(),
            port_profile=port_profile,
            description=str(_cell_value(ws, row, headers.get("description", 4))).strip(),
            access_vlan=_parse_int_cell(access_vlan_raw, "Access VLAN", row, sheet),
            voice_vlan=_parse_int_cell(voice_vlan_raw, "Voice VLAN", row, sheet),
            native_vlan=_parse_int_cell(native_vlan_raw, "Native VLAN", row, sheet),
            allowed_vlans=str(allowed_vlans_raw).strip(),
            storm_control_broadcast=str(storm_control_broadcast_raw).strip(),
            storm_control_multicast=str(storm_control_multicast_raw).strip(),
            port_channel_number=_parse_int_cell(port_channel_raw, "Port Channel No.", row, sheet),
            qos_trust_dscp=qos_trust_dscp,
            template_hint=template_hint,
        ))
    return interfaces


def _build_interface_names(prefix: str, port_start: Any, port_count: Any) -> list[str]:
    if not prefix:
        return []

    start = int(port_start or 1)
    count = int(port_count or 0)
    return [f"{prefix}{port_number}" for port_number in range(start, start + count)]


def _build_unused_interface(
    device_name: str,
    interface_name: str,
    port_profiles: dict[str, Any],
) -> Interface:
    unused_profile = port_profiles.get("unused", {})
    return Interface(
        device_name=device_name,
        interface_name=interface_name,
        port_profile="unused",
        template_hint=str(unused_profile.get("template_hint", "interfaces_unused")),
    )


def _expand_interfaces_from_hardware(
    devices: list[Device],
    interfaces: list[Interface],
    hardware_catalog: dict[str, Any],
    uplink_modules: dict[str, Any],
    port_profiles: dict[str, Any],
) -> list[Interface]:
    if not devices or not hardware_catalog or "unused" not in port_profiles:
        return interfaces

    interfaces_by_device: dict[str, list[Interface]] = {}
    for interface in interfaces:
        interfaces_by_device.setdefault(interface.device_name, []).append(interface)

    expanded: list[Interface] = []
    known_devices = {device.hostname for device in devices}

    for device in devices:
        model_data = hardware_catalog.get(device.model, {})
        module_data = uplink_modules.get(device.uplink_module, {})
        device_interfaces = interfaces_by_device.pop(device.hostname, [])

        inventory = [
            *_build_interface_names(
                str(model_data.get("interface_prefix", "") or ""),
                model_data.get("access_port_start", 1),
                model_data.get("access_ports", 0),
            ),
            *_build_interface_names(
                str(module_data.get("interface_prefix", "") or ""),
                module_data.get("uplink_port_start", 1),
                module_data.get("uplink_ports", 0),
            ),
        ]

        if not inventory:
            expanded.extend(device_interfaces)
            continue

        inventory_names = set(inventory)
        interface_overrides: dict[str, list[Interface]] = {}
        for interface in device_interfaces:
            interface_overrides.setdefault(interface.interface_name, []).append(interface)

        for interface_name in inventory:
            overrides = interface_overrides.get(interface_name)
            if overrides:
                expanded.extend(overrides)
                continue
            expanded.append(_build_unused_interface(device.hostname, interface_name, port_profiles))

        for interface in device_interfaces:
            if interface.interface_name not in inventory_names:
                expanded.append(interface)

    for interface in interfaces:
        if interface.device_name not in known_devices:
            expanded.append(interface)

    return expanded


def _load_global_settings(ws: Worksheet) -> GlobalSettings:
    """Reads key-value pairs from columns A and B."""
    kv: dict[str, Any] = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        if key:
            kv[str(key).strip().lower().replace(" ", "_")] = val
    return GlobalSettings(
        domain_name=str(kv.get("domain_name", "") or "").strip(),
        ntp_servers=_list_field(kv.get("ntp_servers", "")),
        dns_servers=_list_field(kv.get("dns_servers", "")),
        summer_time_config=str(kv.get("summer_time_config", "") or "").strip(),
        aaa_group_name=str(kv.get("aaa_group_name", "TACACS_SERVERS") or "TACACS_SERVERS").strip(),
        tacacs_server_1=str(kv.get("tacacs_server_1", "") or "").strip(),
        tacacs_server_2=str(kv.get("tacacs_server_2", "") or "").strip(),
        tacacs_key=str(kv.get("tacacs_key", "") or "").strip(),
        aaa_fail_message=str(kv.get("aaa_fail_message", "") or "").strip(),
        snmp_auth_protocol=str(kv.get("snmp_auth_protocol", "SHA") or "SHA").strip().upper(),
        snmp_priv_protocol=str(kv.get("snmp_priv_protocol", "AES") or "AES").strip().upper(),
        snmp_ro_group=str(kv.get("snmp_ro_group", "SNMPv3_RO") or "SNMPv3_RO").strip(),
        snmp_ro_user=str(kv.get("snmp_ro_user", "") or "").strip(),
        snmp_ro_auth_password=str(kv.get("snmp_ro_auth_password", "") or "").strip(),
        snmp_ro_priv_password=str(kv.get("snmp_ro_priv_password", "") or "").strip(),
        snmp_rw_group=str(kv.get("snmp_rw_group", "SNMPv3_RW") or "SNMPv3_RW").strip(),
        snmp_rw_user=str(kv.get("snmp_rw_user", "") or "").strip(),
        snmp_rw_auth_password=str(kv.get("snmp_rw_auth_password", "") or "").strip(),
        snmp_rw_priv_password=str(kv.get("snmp_rw_priv_password", "") or "").strip(),
        snmp_host=str(kv.get("snmp_host", "") or "").strip(),
        snmp_location=str(kv.get("snmp_location", "") or "").strip(),
        snmp_contact=str(kv.get("snmp_contact", "") or "").strip(),
        vty_acl=str(kv.get("vty_acl", "ACL_VTY_ACCESS") or "ACL_VTY_ACCESS").strip(),
        snmp_ro_acl=str(kv.get("snmp_ro_acl", "ACL_SNMP_RO_ACCESS") or "ACL_SNMP_RO_ACCESS").strip(),
        snmp_rw_acl=str(kv.get("snmp_rw_acl", "ACL_SNMP_RW_ACCESS") or "ACL_SNMP_RW_ACCESS").strip(),
        syslog_server=str(kv.get("syslog_server", "") or "").strip(),
        banner_motd=str(kv.get("banner_motd", "") or "AUTHORISED ACCESS ONLY. Disconnect immediately if not authorised.").strip(),
        enable_secret=str(kv.get("enable_secret", "changeme") or "changeme").strip(),
        local_username=str(kv.get("local_username", "") or "").strip(),
        local_password=str(kv.get("local_password", "changeme") or "changeme").strip(),
    )


def _load_feature_selection(ws: Worksheet) -> FeatureSelection:
    """Reads feature name (col A) and enabled Yes/No (col B)."""
    kv: dict[str, bool] = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        if name:
            kv[str(name).strip().lower().replace(" ", "_")] = str(val).strip().lower() in ("yes", "true", "1")
    return FeatureSelection(
        base_config=kv.get("base_config", True),
        vlans=kv.get("vlans", True),
        interfaces=kv.get("interfaces", True),
        acls=kv.get("acls", True),
    )


def _load_acls(ws: Worksheet) -> list[ACLEntry]:
    """Reads ACL entries from the ACLs sheet."""
    headers = _header_map(ws)
    _validate_required_headers(headers, ["acl_name"], ws.title)
    entries: list[ACLEntry] = []
    for row in range(2, ws.max_row + 1):
        acl_name = _cell_value(ws, row, headers.get("acl_name", 1))
        if not acl_name:
            continue
        entries.append(ACLEntry(
            acl_name=str(acl_name).strip(),
            remark=str(_cell_value(ws, row, headers.get("remark", 2))).strip(),
            action=str(_cell_value(ws, row, headers.get("action", 3))).strip().lower(),
            network=str(_cell_value(ws, row, headers.get("network/host", 4))).strip(),
            wildcard=str(_cell_value(ws, row, headers.get("wildcard", 5))).strip(),
        ))
    return entries


def load_workbook(
    path: str | Path,
    port_profiles: dict[str, Any] | None = None,
    hardware_catalog: dict[str, Any] | None = None,
    uplink_modules: dict[str, Any] | None = None,
) -> Intent:
    """Parse an intent workbook and return an Intent object."""
    wb: Workbook = openpyxl.load_workbook(str(path), data_only=True)

    def _get_sheet(name: str) -> Worksheet | None:
        for sname in wb.sheetnames:
            if sname.lower() == name:
                return wb[sname]
        return None

    devices_ws = _get_sheet("devices")
    vlans_ws = _get_sheet("vlans")
    interfaces_ws = _get_sheet("interfaces")
    global_ws = _get_sheet("global settings")
    features_ws = _get_sheet("feature selection")
    acls_ws = _get_sheet("acls")

    devices = _load_devices(devices_ws) if devices_ws else []
    loaded_interfaces = _load_interfaces(interfaces_ws, port_profiles or {}) if interfaces_ws else []
    interfaces = _expand_interfaces_from_hardware(
        devices=devices,
        interfaces=loaded_interfaces,
        hardware_catalog=hardware_catalog or {},
        uplink_modules=uplink_modules or {},
        port_profiles=port_profiles or {},
    )

    return Intent(
        devices=devices,
        vlans=_load_vlans(vlans_ws) if vlans_ws else [],
        interfaces=interfaces,
        global_settings=_load_global_settings(global_ws) if global_ws else GlobalSettings(),
        feature_selection=_load_feature_selection(features_ws) if features_ws else FeatureSelection(),
        acls=_load_acls(acls_ws) if acls_ws else [],
    )
